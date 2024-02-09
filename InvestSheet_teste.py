import pandas as pd
from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
import re
from sys import exit

acs = []

# criar uma __str__() para adicionar a ação, que você acabou de digitar, na predefinição
# tipo: if input("voce quer colocar essa ação na predefinição?"), mod.write(acao_instancia)


class Acoes:

    @classmethod
    # não está funcionando, o "pattern" está correto só que não termina o "while"
    def pegar_nome(cls):

        while True:
            nome = str(
                input("\nQual ação você quer colocar em sua planilha?\n")).strip()
            if not nome:
                print("\nSem nome de ação identificado, digite novamente.")

            # if re.search(r"^......?$", nome, re.IGNORECASE): #^[a-z]{4}[1-9]{1-2}$
                # break
            # else:
                # print("\nNome inválido, digite novamente")
            if nome:
                print(nome)
                break

        return nome

    @classmethod
    # ainda não foi testado para ver se funciona
    def pegar_coluna(cls):
        while True:
            coluna = input("Qual ação você quer colocar em sua planilha?\n")
            if not coluna:
                print("\nSem coluna identificada, digite novamente.")

            try:
                str(coluna).strip().upper()
            except ValueError:
                print("\nO nome da coluna não é aceito, digite novamente.")
                continue

            if tab2[f'{coluna}3'].value == None:
                break

            else:
                while True:
                    sob = input(
                        "Esta coluna já está ocupada, você quer escrever sobre o conteúdo da coluna? (isso o deletará)\n"
                        "[1] Sim\n[2] Não\n"
                    )

                    if sob == 1 or sob == 2:
                        break

                    else:
                        print("Valor de resposta inadequado, digite novamente")

            if sob == 1:
                break

        return coluna
    # ainda não foi testado para ver se funciona

    def __init__(self, nome, coluna):
        self.nome = nome
        self.coluna = coluna
    # ainda não foi testado para ver se funciona

    def procurar_ac(self):
        chrome = webdriver.Chrome()
        chrome.get(f'https://statusinvest.com.br/acoes/{self.nome}')
        preco = chrome.find_element(
            By.XPATH, '//*[@id="main-2"]/div[2]/div/div[1]/div/div[1]/div/div[1]/strong').text
        roe = chrome.find_element(
            By.XPATH, '//*[@id="indicators-section"]/div[2]/div/div[4]/div/div[1]/div/div/strong').text
        dy = chrome.find_element(
            By.XPATH, '//*[@id="main-2"]/div[2]/div/div[1]/div/div[4]/div/div[1]/strong').text
        div_ebitda = chrome.find_element(
            By.XPATH, '//*[@id="indicators-section"]/div[2]/div/div[2]/div/div[2]/div/div/strong').text
        pl = chrome.find_element(
            By.XPATH, '//*[@id="indicators-section"]/div[2]/div/div[1]/div/div[2]/div/div/strong').text
        pvp = chrome.find_element(
            By.XPATH, '//*[@id="indicators-section"]/div[2]/div/div[1]/div/div[4]/div/div/strong').text
        payout = chrome.find_element(
            By.XPATH, '//*[@id="payout-section"]/div/div/div[1]/div[1]/div[1]/strong').text
        cagrL = chrome.find_element(
            By.XPATH, '//*[@id="indicators-section"]/div[2]/div/div[5]/div/div[2]/div/div/strong').text
        cagrR = chrome.find_element(
            By.XPATH, '//*[@id="payout-section"]/div/div/div[1]/div[1]/div[1]/strong').text

        # o número não conseguia ser convertido em float, pois tinha vírgula, e o float utiliza na numeração americana
        tab2[f'{self.coluna}1'] = str(self.nome.upper())
        tab2[f'{self.coluna}3'] = float(preco.replace(",", "."))
        try:
            tab2[f'{self.coluna}4'] = float(
                roe.replace("%", "").replace(",", "."))
        except ValueError:
            print("\nO valor do ROE não está disponível.")
            tab2[f'{self.coluna}4'] = "-"

        try:
            tab2[f'{self.coluna}5'] = float(dy.replace(",", "."))
        except ValueError:
            print("\nO valor de Dividend Yield não está disponível.")
            tab2[f'{self.coluna}5'] = "-"

        try:
            tab2[f'{self.coluna}6'] = float(div_ebitda.replace(",", "."))
        except ValueError:
            print("\nO valor de Dívida Líquida/ EBITDA não está disponível.")
            tab2[f'{self.coluna}6'] = "-"

        try:
            tab2[f'{self.coluna}7'] = float(pl.replace(",", "."))
        except:
            print("\nO valor do PL não está disponível.")
            tab2[f'{self.coluna}7'] = "-"

        try:
            tab2[f'{self.coluna}8'] = float(pvp.replace(",", "."))
        except ValueError:
            print("\nO valor do PVP não está disponível.")
            tab2[f'{self.coluna}8'] = "-"

        try:
            tab2[f'{self.coluna}9'] = float(
                payout.replace("%", ".").replace(",", "."))
        except ValueError:
            print("\nO valor do PAYOUT não está disponível.")
            tab2[f'{self.coluna}9'] = "-"

        try:
            tab2[f'{self.coluna}10'] = float(
                cagrL.replace("%", "").replace(",", "."))
        except ValueError:
            print("\nO valor de CAGR do Lucro não está disponível.")
            tab2[f'{self.coluna}10'] = "-"

        try:
            tab2[f'{self.coluna}11'] = float(
                cagrR.replace("%", "").replace(",", "."))
        except ValueError:
            print("\nO valor de CAGR dee Receita não está disponível.")
            tab2[f'{self.coluna}11'] = "-"

    # __str__ é utilizado no main1() nas observações
    def __str__(self):
        return f"{self.coluna.capitalize().strip()};{self.nome.strip()}"


def atualizar_predef():
    global acs
    with open("moderador.txt", "r") as mod:
        for i in mod:
            coluna, acao = i.split(";")
            acs.append({
                "col": coluna,
                "ac": acao,
            })


def mostrar_predef():
    print("Sua predefinição de ações:\n")
    for a in acs:
        print(a["col"], a["ac"], sep=", ")


def adicionar_predef():
    ac = input("Você quer adicionar qual ação?\n")
    col = input("\nEm qual coluna está essa ação?\n")
    with open("moderador.txt", "a") as mod:
        mod.write(f"{col.capitalize().strip()};{ac.strip()}")


def main1():
    global acao_instancia
    ac_nome = Acoes.pegar_nome()
    ac_coluna = Acoes.pegar_coluna()
    acao_instancia = Acoes(ac_nome, ac_coluna)
    acao_instancia.procurar_ac()
    # esc2 = input("Você quer adicionar a ação na predefinição?\n[1] Sim\n[2] Não\n")
    # if esc2 == 2:
    #    with open("moderador.txt", "a") as mod:
    #        mod.write(acao_instancia)


def main2():
    atualizar_predef()
    while True:
        p1 = int(input(
            "[1] Mostrar predefinição\n[2] Adicionar ação\n[3] Preencher planilha\n[4] Sair\n"))
        if p1 == 1:
            mostrar_predef()
        if p1 == 2:
            adicionar_predef()
            atualizar_predef()
        if p1 == 3:
            for n in acs:
                Acoes.procurar_ac(n["col"], n["ac"])
        if p1 == 4:
            exit()


while True:
    nome = int(
        input("Qual o nome da planilha?\n[1] TABELA DE AÇÕES\n[2] Outro\n"))

    if nome == 1:
        tabela = load_workbook('TABELA DE AÇÕES.xlsx')

    if nome == 2:
        nome2 = input("Qual o nome da planilha?\n")
        tabela = load_workbook(f'{nome2}.xlsx')

    tab2 = tabela.active

    esc1 = int(
        input("[1] Adicionar ações manualmente\n[2] Utilizar a predefinição\n"))

    if esc1 == 1:
        main1()
    elif esc1 == 2:
        main2()

    if nome == 1:
        tabela.save('TABELA DE AÇÕES.xlsx')
    else:
        tabela.save(f'{nome2}.xlsx')

    while True:
        t = int(input("Você quer colocar outra ação?\n[1] Sim\n[2] Não\n"))
        if t == 2 or t == 1:
            break
        else:
            continue
    if t == 2:
        break
